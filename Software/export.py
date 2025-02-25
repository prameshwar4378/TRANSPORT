import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from .filters import *
from openpyxl.styles import Font 
from datetime import date
from .filters import *
from Developer.models import *
from django.utils.timezone import localtime
from django.db.models import Sum, Count, Q, F

def export_filtered_bill_records(request):
    queryset = Bill.objects.all().order_by('id')
    filter = BillFilter(request.GET, queryset=queryset)
    filtered_rec = filter.qs  # Filtered queryset

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Report"

    # Define the header row for the Excel sheet
    headers = ['Bill Number', 'Vehicle Number', 'Vehicle Owner', 'Driver', 'Party', 
               'From Location', 'To Location', 'Commission Amount', 'Commission Advance',  'Commission Pending', 
               'Commission Pay Date', 'Bill Date']
    ws.append(headers)

    # Loop through each filtered record and append data to the sheet
    for bill in filtered_rec:
        row = [
            bill.bill_number, 
            bill.vehicle.vehicle_number if bill.vehicle else '',  # Assuming 'vehicle_number' is a field in the 'Vehicle' model
            bill.vehicle.owner.owner_name if bill.vehicle.owner.owner_name else '',  # Assuming 'name' is the field you want from the VehicleOwner model
            bill.driver.driver_name if bill.driver else '',  # Assuming 'name' is the field you want from the Driver model
            bill.party.name if bill.party else '',  # Assuming 'name' is the field you want from the Party model
            bill.from_location,
            bill.to_location,
            bill.commission_charge,
            bill.commission_received,  
            bill.commission_pending,  
            bill.commission_received_date,  # Assuming commission_received_date is the commission pay date
            bill.bill_date,
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Bill Details.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response



def export_todays_bill_records(request):
    # Filter bills with today's date
    today = date.today()
    queryset = Bill.objects.filter(bill_date=today).order_by('id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Report"

    # Define the header row for the Excel sheet
    headers = ['Bill Number', 'Vehicle Number', 'Vehicle Owner', 'Driver', 'Party', 
               'From Location', 'To Location', 'Commission Amount', 'Commission Advance',  
               'Commission Pending', 'Commission Pay Date', 'Bill Date']
    ws.append(headers)

    # Loop through each filtered record and append data to the sheet
    for bill in queryset:
        row = [
            bill.bill_number, 
            bill.vehicle.vehicle_number if bill.vehicle else '',  
            bill.vehicle.owner.owner_name if bill.vehicle and bill.vehicle.owner else '',  
            bill.driver.driver_name if bill.driver else '',  
            bill.party.name if bill.party else '',  
            bill.from_location,
            bill.to_location,
            bill.commission_charge,
            bill.commission_received,  
            bill.commission_pending,  
            bill.commission_received_date.strftime("%Y-%m-%d") if bill.commission_received_date else '',  
            bill.bill_date.strftime("%Y-%m-%d") if bill.bill_date else '',
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Bill_Details_{today}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response



def export_bill_records_for_single_owner(request,id):
    # Filter bills with today's date
    owner=VehicleOwner.objects.get(id=id)
    today = date.today()
    queryset = Bill.objects.filter(vehicle__owner=owner).order_by('-id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Report"

    # Define the header row for the Excel sheet
    headers = ['Bill Number', 'Vehicle Number',  'Driver', 'Party', 
               'From Location', 'To Location', 'Commission Amount', 'Commission Advance',  
               'Commission Pending', 'Commission Pay Date', 'Bill Date']
    ws.append(headers)

    # Loop through each filtered record and append data to the sheet
    for bill in queryset:
        row = [
            bill.bill_number, 
            bill.vehicle.vehicle_number if bill.vehicle else '',  
            bill.driver.driver_name if bill.driver else '',  
            bill.party.name if bill.party else '',  
            bill.from_location,
            bill.to_location,
            bill.commission_charge,
            bill.commission_received,  
            bill.commission_pending,  
            bill.commission_received_date.strftime("%Y-%m-%d") if bill.commission_received_date else '',  
            bill.bill_date.strftime("%Y-%m-%d") if bill.bill_date else '',
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={owner.owner_name} - {today}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response




def export_bill_records_for_single_vehicle(request,id):
    # Filter bills with today's date
    vehicle=Vehicle.objects.get(id=id)
    today = date.today()
    queryset = Bill.objects.filter(vehicle=vehicle).order_by('-id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Report"

    # Define the header row for the Excel sheet
    headers = ['Bill Number', 'Vehicle Number',  'Vehicle Owner',  'Driver', 'Party', 
               'From Location', 'To Location', 'Commission Amount', 'Commission Advance',  
               'Commission Pending', 'Commission Pay Date', 'Bill Date']
    ws.append(headers)

    # Loop through each filtered record and append data to the sheet
    for bill in queryset:
        row = [
            bill.bill_number, 
            bill.vehicle.vehicle_number if bill.vehicle else '',  
            bill.vehicle.owner.owner_name if bill.vehicle and bill.vehicle.owner else '',  
            bill.driver.driver_name if bill.driver else '',  
            bill.party.name if bill.party else '',  
            bill.from_location,
            bill.to_location,
            bill.commission_charge,
            bill.commission_received,  
            bill.commission_pending,  
            bill.commission_received_date.strftime("%Y-%m-%d") if bill.commission_received_date else '',  
            bill.bill_date.strftime("%Y-%m-%d") if bill.bill_date else '',
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={vehicle.vehicle_number} - {today}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response



def export_bill_records_for_single_party(request,id):
    # Filter bills with today's date
    party=Party.objects.get(id=id)
    today = date.today()
    queryset = Bill.objects.filter(party=party).order_by('-id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Report"

    # Define the header row for the Excel sheet
    headers = ['Bill Number', 'Vehicle Number',  'Vehicle Owner',  'Driver', 'Party', 
               'From Location', 'To Location', 'Commission Amount', 'Commission Advance',  
               'Commission Pending', 'Commission Pay Date', 'Bill Date']
    ws.append(headers)

    # Loop through each filtered record and append data to the sheet
    for bill in queryset:
        row = [
            bill.bill_number, 
            bill.vehicle.vehicle_number if bill.vehicle else '',  
            bill.vehicle.owner.owner_name if bill.vehicle and bill.vehicle.owner else '',  
            bill.driver.driver_name if bill.driver else '',  
            bill.party.name if bill.party else '',  
            bill.from_location,
            bill.to_location,
            bill.commission_charge,
            bill.commission_received,  
            bill.commission_pending,  
            bill.commission_received_date.strftime("%Y-%m-%d") if bill.commission_received_date else '',  
            bill.bill_date.strftime("%Y-%m-%d") if bill.bill_date else '',
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={party.name} - {today}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response




def export_bill_records_for_single_driver(request,id):
    # Filter bills with today's date
    driver=Driver.objects.get(id=id)
    today = date.today()
    queryset = Bill.objects.filter(driver=driver).order_by('-id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill Report"

    # Define the header row for the Excel sheet
    headers = ['Bill Number', 'Vehicle Number',  'Vehicle Owner',  'Driver', 'Party', 
               'From Location', 'To Location', 'Commission Amount', 'Commission Advance',  
               'Commission Pending', 'Commission Pay Date', 'Bill Date']
    ws.append(headers)

    # Loop through each filtered record and append data to the sheet
    for bill in queryset:
        row = [
            bill.bill_number, 
            bill.vehicle.vehicle_number if bill.vehicle else '',  
            bill.vehicle.owner.owner_name if bill.vehicle and bill.vehicle.owner else '',  
            bill.driver.driver_name if bill.driver else '',  
            bill.party.name if bill.party else '',  
            bill.from_location,
            bill.to_location,
            bill.commission_charge,
            bill.commission_received,  
            bill.commission_pending,  
            bill.commission_received_date.strftime("%Y-%m-%d") if bill.commission_received_date else '',  
            bill.bill_date.strftime("%Y-%m-%d") if bill.bill_date else '',
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={driver.driver_name} - {today}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


